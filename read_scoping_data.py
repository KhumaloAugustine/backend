import openpyxl
import json

# Load the scoping file
path = r'c:\Users\Augustine.Khumalo\Documents\PAMHoYA\backend\Scoping for mental health data\Scoping - Mental Health Datasets in SA.xlsx'
wb = openpyxl.load_workbook(path)
print('Sheet names:', wb.sheetnames)

# Get first sheet
ws = wb.active
print('\nFirst 20 rows of', ws.title, ':')
for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
    if i < 20:
        print(f'Row {i}:', [str(v)[:40] if v else '' for v in row[:8]])

# Also check supplementary file
path2 = r'c:\Users\Augustine.Khumalo\Documents\PAMHoYA\backend\Scoping for mental health data\Suplementarry tables 11 May 2025.xlsx'
wb2 = openpyxl.load_workbook(path2)
print('\n\nSheet names in supplementary:', wb2.sheetnames)
ws2 = wb2.active
print('First 10 rows of', ws2.title, ':')
for i, row in enumerate(ws2.iter_rows(max_row=10, values_only=True)):
    if i < 10:
        print(f'Row {i}:', [str(v)[:40] if v else '' for v in row[:6]])
