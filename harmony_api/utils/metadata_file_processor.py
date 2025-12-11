"""
MIT License

Copyright (c) 2025 PAMHoYA Team (https://pamhoya.ac.uk)
Project: PAMHoYA - Platform for Advancing Mental Health in Youth and Adolescence

Utility functions for metadata processing from various file formats.
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
import re

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)


class MetadataFileProcessor:
    """Utility class for processing metadata from various file formats"""

    @staticmethod
    def extract_from_json(filepath: str) -> Optional[Dict[str, Any]]:
        """
        Extract metadata from JSON file.

        Args:
            filepath: Path to JSON file

        Returns:
            Parsed JSON dictionary or None if error
        """
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in {filepath}: {str(e)}")
            return None
        except Exception as e:
            logger.error(f"Error reading {filepath}: {str(e)}")
            return None

    @staticmethod
    def extract_from_pdf(filepath: str) -> Optional[Dict[str, Any]]:
        """
        Extract basic metadata from PDF file.

        Args:
            filepath: Path to PDF file

        Returns:
            Dictionary with extracted metadata or None if error
        """
        if not PdfReader:
            logger.warning("PyPDF2 not installed. Cannot process PDF files.")
            return None

        try:
            with open(filepath, "rb") as f:
                pdf = PdfReader(f)

                # Extract metadata from PDF properties
                metadata = {
                    "source_name": Path(filepath).stem,
                    "pages": len(pdf.pages),
                }

                # Try to extract document info
                if pdf.metadata:
                    if pdf.metadata.title:
                        metadata["title"] = pdf.metadata.title
                    if pdf.metadata.author:
                        metadata["authors"] = [pdf.metadata.author]
                    if pdf.metadata.subject:
                        metadata["abstract"] = pdf.metadata.subject
                    if pdf.metadata.producer:
                        metadata["producer"] = pdf.metadata.producer

                return metadata

        except Exception as e:
            logger.error(f"Error reading PDF {filepath}: {str(e)}")
            return None

    @staticmethod
    def extract_from_excel(filepath: str) -> Optional[List[Dict[str, Any]]]:
        """
        Extract metadata from Excel file (questionnaires metadata).

        Args:
            filepath: Path to Excel file

        Returns:
            List of row dictionaries or None if error
        """
        if not openpyxl:
            logger.warning("openpyxl not installed. Cannot process Excel files.")
            return None

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            # Extract headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Extract data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = cell.value
                if any(row_data.values()):  # Skip empty rows
                    data.append(row_data)

            return data if data else None

        except Exception as e:
            logger.error(f"Error reading Excel {filepath}: {str(e)}")
            return None

    @staticmethod
    def extract_from_text(filepath: str) -> Optional[Dict[str, Any]]:
        """
        Extract basic metadata from text file.

        Args:
            filepath: Path to text file

        Returns:
            Dictionary with extracted metadata or None if error
        """
        try:
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()

            metadata = {
                "source_name": Path(filepath).stem,
                "file_size": len(content),
                "lines": content.count("\n"),
            }

            # Try to extract title (first non-empty line)
            lines = [line.strip() for line in content.split("\n") if line.strip()]
            if lines:
                metadata["title"] = lines[0][:200]  # First 200 chars

            return metadata

        except Exception as e:
            logger.error(f"Error reading text file {filepath}: {str(e)}")
            return None

    @staticmethod
    def guess_metadata_format(content: str) -> str:
        """
        Guess the format of metadata content.

        Args:
            content: Content string to analyze

        Returns:
            Format string: 'json', 'xml', 'ddi', or 'unknown'
        """
        content_lower = content.strip().lower()

        if content_lower.startswith("{") or content_lower.startswith("["):
            return "json"
        elif content_lower.startswith("<"):
            return "xml"
        elif "codeBook" in content or "ddi:codeBook" in content:
            return "ddi"
        else:
            return "unknown"

    @staticmethod
    def normalize_metadata_keys(metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Normalize metadata keys to standard format.

        Args:
            metadata: Dictionary with metadata

        Returns:
            Dictionary with normalized keys
        """
        normalized = {}

        # Common key mappings
        key_mappings = {
            "title": ["title", "name", "study_title", "dataset_title"],
            "abstract": ["abstract", "description", "summary"],
            "keywords": ["keywords", "tags", "topics"],
            "source_name": ["source_name", "study_name", "organization"],
            "authors": ["authors", "creators", "producers"],
            "date_created": [
                "date_created",
                "created_date",
                "start_date",
                "start_time",
            ],
            "date_updated": ["date_updated", "updated_date", "end_date", "end_time"],
            "version": ["version", "version_number"],
            "license": ["license", "license_info", "rights"],
        }

        # Try to find matching keys
        for standard_key, possible_keys in key_mappings.items():
            for key, value in metadata.items():
                if key.lower() in [k.lower() for k in possible_keys]:
                    normalized[standard_key] = value
                    break

        # Keep any unmapped keys
        for key, value in metadata.items():
            if key not in normalized:
                normalized[key] = value

        return normalized

    @staticmethod
    def extract_email_addresses(text: str) -> List[str]:
        """Extract email addresses from text"""
        pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
        return re.findall(pattern, text)

    @staticmethod
    def extract_urls(text: str) -> List[str]:
        """Extract URLs from text"""
        pattern = r"https?://[^\s]+"
        return re.findall(pattern, text)

    @staticmethod
    def extract_dois(text: str) -> List[str]:
        """Extract DOI references from text"""
        pattern = r"https://doi\.org/[^\s]+"
        return re.findall(pattern, text)


class QuestionnairesMetadataProcessor:
    """Specialized processor for questionnaire metadata files"""

    @staticmethod
    def extract_from_xlsx_questionnaires(filepath: str) -> Optional[Dict[str, Any]]:
        """
        Extract questionnaire metadata from Excel file.

        Expected format:
        - Column A: Questionnaire Name
        - Column B: Study/Dataset
        - Column C: Items/Questions
        - Column D: Response Options
        - Additional columns for language, version, etc.

        Args:
            filepath: Path to Excel file

        Returns:
            Dictionary with questionnaire metadata
        """
        processor = MetadataFileProcessor()
        data = processor.extract_from_excel(filepath)

        if not data:
            return None

        questionnaires = {
            "source_file": filepath,
            "questionnaires": [],
        }

        for row in data:
            questionnaire = {
                "name": row.get("Questionnaire Name") or row.get("Name"),
                "study": row.get("Study") or row.get("Dataset"),
                "items_count": row.get("Items") or row.get("Questions"),
                "response_options": row.get("Response Options"),
                "language": row.get("Language", "English"),
                "version": row.get("Version"),
                "raw_data": row,
            }
            questionnaires["questionnaires"].append(questionnaire)

        return questionnaires

    @staticmethod
    def extract_from_pdf_questionnaires(pdf_files: List[str]) -> Dict[str, Any]:
        """
        Extract metadata from multiple PDF questionnaire files.

        Args:
            pdf_files: List of paths to PDF questionnaire files

        Returns:
            Dictionary with aggregated questionnaire metadata
        """
        processor = MetadataFileProcessor()
        questionnaires = {
            "source_files": pdf_files,
            "questionnaires": [],
        }

        for pdf_file in pdf_files:
            metadata = processor.extract_from_pdf(pdf_file)
            if metadata:
                # Try to extract language from filename
                filename = Path(pdf_file).stem
                language = "Unknown"
                if "Xhosa" in filename or "Xhosa" in filename:
                    language = "IsiXhosa"
                elif "Zulu" in filename:
                    language = "isiZulu"
                elif "Sotho" in filename or "Sepedi" in filename:
                    language = "Sesotho/Sepedi"
                elif "Tswana" in filename:
                    language = "Setswana"

                metadata["language"] = language
                questionnaires["questionnaires"].append(metadata)

        return questionnaires
